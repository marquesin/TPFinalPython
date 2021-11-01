import time
from openpyxl import Workbook
from openpyxl import load_workbook
from os import path


# Si no existe workbook lo inicia, si no lo carga
if path.exists('registros.xlsx') == False:
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Cliente'
    ws['B1'] = 'Fecha'
    ws['C1'] = 'Combo S'
    ws['D1'] = 'Combo D'
    ws['E1'] = 'Combo T'
    ws['F1'] = 'Flurby'
    ws['G1'] = 'Total'
else:
    wb = load_workbook('registros.xlsx')
    ws = wb.active


def excel():
    """Función que guarda info en el excel"""
    ws.append([nombre_cl, fecha, simple, doble, triple, postre, total])

    wb.save("registros.xlsx")
    return print('Excel guardado')


def ingreso():
    """Guarda ingreso de encargado en el txt"""
    fecha = time.asctime()
    f = open("registros.txt", "a")
    f.write("IN " + fecha + " " + "Encargad@ " + nombre + "\n")
    f.close()
    print("Registro salvado")


def salida():
    """Guarda salida del encargado en el txt"""
    fecha = time.asctime()
    f = open("registros.txt", "a")
    f.write("OUT " + fecha + " " + "Encargad@ " +
            nombre + " $" + str(totalturno)+"\n")
    f.close()
    print("Registro salvado")


def tienenumero(campo):
    """Itera el string, devuelve True si hay un número"""
    return any(letra.isdigit() for letra in campo)


def verificar(campo):
    """Verifica que no haya números en un string, de lo contrario lo vuelve a pedir"""

    while campo == " " or campo.isdecimal() == True or tienenumero(campo) == True:
        print("Error. El dato ingresado es incorrecto.\nIngrese solo letras.")
        campo = input("Intente de nuevo: ")
    return campo


def verificar_nro(campo):
    """Verifica si el elemento es un número, de ser así lo devuelve como int, de lo contrario lo vuelve a pedir."""

    while campo == " " or campo.isdecimal() == False or tienenumero(campo) == False:
        print("Error. El dato ingresado es incorrecto.\nIngrese solo números.")
        campo = input("Intente de nuevo: ")
    return int(campo)


def nuevo_pedido():
    """Contiene los valores de cada combo y pide la cantidad de cada uno como inputs. 
    Devuelve una lista con total y cantidades en forma individual"""
    # precios
    combo_s = 650
    combo_d = 700
    combo_t = 800
    postre = 250

    # cantidades pedidas
    mult_s = verificar_nro(input('Ingrese cantidad Combo S: '))
    mult_d = verificar_nro(input('Ingrese cantidad Combo D: '))
    mult_t = verificar_nro(input('Ingrese cantidad Combo T: '))
    mult_p = verificar_nro(input('Ingrese cantidad Flurby: '))

    # calculo del total
    total = (combo_s*mult_s + combo_d * mult_d +
             combo_t*mult_t + postre * mult_p)

    return [total, mult_s, mult_d, mult_t, mult_p]


# inicializamos nombre, lo ingresamos al txt e inicializamos el total del turno.
nombre = verificar(input('\nIngrese su nombre encargad@: ').title())
ingreso()
totalturno = 0

# menú
while True:
    print(f'{" ":20}McDowell\'s{" ":20}\n')
    print('Recuerda que siempre hay que recibir al cliente con una sonrisa :)\n')
    print(f'{" ":10}1 - Ingreso de nuevo pedido{" ":10}')
    print(f'{" ":10}2 - Cambio de turno{" ":10}')
    print(f'{" ":10}3 - Apagar sistema{" ":10}\n')
    opcion = input('Seleccione una opción: ')

    if opcion == '1':
        fecha = time.asctime()

        # nombre de cliente
        nombre_cl = verificar(
            input('\nIngrese el nombre del cliente: ').title())
        pedido = nuevo_pedido()  # devuelve una lista

        # unpackeamos la lista en estas variables para poder pasarlas como argumentos de openpyxl
        total, simple, doble, triple, postre = pedido

        print(f'\n{" ":>30}Total: ${total}')
        pago = verificar_nro(input('Abona con: '))

       # check para ver si el pago es suficiente
        while pago < total:
            print(f'ERROR: El pago debería ser igual o mayor al total')
            pago = verificar_nro(input('Abona con: '))

        # calcula vuelto
        vuelto = (pago - total)
        print(f'\n{" ":>30}Vuelto: ${vuelto}\n')
        print(f'{" ":10}---------------------------------{" ":10}\n')

        excel()
        # agrega el total al turno actual
        totalturno += total

    elif opcion == '2':

        salida()
        nombre = verificar(input(f'\nIngrese su nombre encargad@: ').title())
        totalturno = 0
        ingreso()

    elif opcion == '3':
        salida()
        break

    else:
        print(f'ERROR: Elija una opción válida\n')
